from audioop import add
from cmath import exp
from datetime import timezone
from importlib.resources import read_text
from multiprocessing import current_process
from pprint import pprint
from traceback import print_tb
from typing_extensions import get_origin
from numpy import fabs
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import time
import os
import requests
import json
from threading import Timer
from matplotlib import pyplot as plt
plt.switch_backend('agg')
plt.rcParams.update({'figure.max_open_warning': 0})
import pandas as pd
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 130)
pd.set_option('display.precision', 3)
import seaborn as sb
import midas.utils as utils
import fm_settings
import serial_number_parser
from openpyxl import load_workbook

def get_current_time():
    current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())
    return current_time
# print("current time is " + get_current_time())
    
FILENAMES = {
    'cp3030': 'Test_summary_30deg.tsv',
    'cp6060': 'Test_summary_60deg.tsv',
    'cv3060': 'Test_summary_A30S60deg.tsv',
    'cv2237': 'Test_summary_A22S37Verification.tsv',
    'vis2237': 'Test_summary_A22S37Visual.tsv',
    'prefix': 'Test_summary_A',
    'verifi': 'Verification.tsv',
    'visual': 'Visual.tsv',
    'focus': 'Test_summary_Focus.tsv'
    }
    
folder_to_track = fm_settings.path_to_track
countdown_time = fm_settings.countdown_time_set
time_out = fm_settings.time_out
script_fm_path = fm_settings.script_path
url = fm_settings.url
url_calibration = fm_settings.url + r'/calibration'
url_focus = fm_settings.url + r'/focus'
url_module_log = fm_settings.url + r'/module_log'
url_mpi = fm_settings.url + r'/mpi'
url_error = fm_settings.url + r'/error'

bool_modified = False
bool_parsing_data = False
bool_changes = False
data0 = ""
new_folder_count = 1

print(folder_to_track)
print(countdown_time)
print(time_out)
print(script_fm_path)

# Create a json file contains pointer informations
try:
    with open(script_fm_path + "pointers.json") as file:
        pointers = json.load(file)
except:
    with open(script_fm_path + "pointers.json","w+") as file:
        pointers = {}
        pointers['Example']={
            'Path': os.path.dirname(os.path.abspath(__file__)),
            'Pending to Parse': False,
            'Rows': 0,
            'Pointer': 0,
            'Last Update': get_current_time(),
            'Days Left': 999999,
            'Existence': True
            }
        json.dump(pointers, file, indent=4)


def create_pointer(path):
    with open(script_fm_path + "pointers.json","r") as file:
        pointers = json.load(file)
    with open(script_fm_path + "pointers.json","w+") as file:
        try:
            row_num = sum(1 for line in open(path)) - 2
            
        except:
            wb = load_workbook(path, read_only=True)
            sheet = wb.worksheets[0]
            row_num = sheet.max_row - 2
            print(create_pointer()+"create_pointers_expect")
            
        pointers[path]={
            'Path': os.path.dirname(path),
            'Pending to Parse': True,
            'Rows': row_num,
            'Pointer': -1,
            'Last Update': get_current_time(),
            'Days Left': 14,
            'Existence': True
            }
        json.dump(pointers, file, indent=4)
        print("finish create_pointer"+create_pointer())

def update_pointer(path, newPointer):
    global bool_changes
    with open(script_fm_path + "pointers.json","r") as file:
        pointers = json.load(file)
        data = pointers[path]
    try:
            row_num = sum(1 for line in open(path)) - 2
            #print()
    except:
        wb = load_workbook(path, read_only=True)
        sheet = wb.worksheets[0]
        row_num = sheet.max_row - 2
    data['Rows'] = row_num
    if data['Rows'] != data['Pointer']:
        data['Pending to Parse'] = True
        data['Last Update'] = get_current_time()
        data['Days Left'] = 14
        bool_changes = True
        # print()

    if newPointer >= 0:
        data['Pointer'] = newPointer
        # print()

    pointers[path] = data
    with open(script_fm_path + "pointers.json","w+") as file:
        json.dump(pointers, file, indent=4)
    # print()


def new_day():
    with open(script_fm_path + "pointers.json", "r") as file:
        pointers = json.load(file)

    for path in pointers:
        with open(script_fm_path + "pointers.json", "r") as file:
            data = pointers[path]
        data['Days Left'] -= 1
        with open(script_fm_path + "pointers.json","w+") as file:
            json.dump(pointers, file, indent=4)

def delete_outdated_pointer():
    with open(script_fm_path + "pointers.json", "r") as file:
        pointers = json.load(file)

    # Check Days Left in json file
    Checking = True
    while Checking:
        Checking = False
        try:
            for path in pointers:
                with open(script_fm_path + "pointers.json", "r") as file:
                    data = pointers[path]

                if data['Days Left'] < 0 or data['Existence'] == False:
                    del pointers[path]
                    with open(script_fm_path + "pointers.json","w+") as file:
                        json.dump(pointers, file, indent=4)
                    Checking = True
        except:
            pass

def upload_log(path, text): # Save upload_log.txt File
    log_file_path = path + r"/Upload_Log.txt"
    try:
        with open(log_file_path,"a") as file:
            file.write(text + "\n")
    except:
        print("error in writing upload_log.txt")
        NAS_fm_log("error in writing upload_log.txt")

def NAS_fm_log(text): # Save NAS_fm_log.txt File
    log_file_path = script_fm_path + r"/NAS_fm_log.txt"
    try:
        with open(log_file_path,"a") as file:
            file.write("Time:" + get_current_time() + "\t" + text + "\n")
    except:
        print("error in writing NAS_fm_log.txt")

def load_data(path, file_name):
    global data0
    data0 = utils.CalibrationData([path], file_name, uniquify=False, typify=False, verbose=False)

def fm_main():
    global bool_parsing_data, bool_modified, bool_changes, countdown_time

    countdown_time = fm_settings.countdown_time_set
    # print("+++++++++++++++++++++++++++++++++++++++")
    if bool_modified == False and bool_changes == True:
        bool_modified = True
        bool_changes = False
        while countdown_time >= 0:
            print(f'Start Parsing Data After {str(countdown_time).zfill(2)} s.\r', end="")
            NAS_fm_log(f'Start Parsing Data After {str(countdown_time).zfill(2)} s.')
            time.sleep(1)
            countdown_time -= 1
        
        bool_modified = False
        bool_parsing_data = True

        # Start Parsing
        print("\nStart to Parse the Data:")
        NAS_fm_log("Start to Parse the Data:")
        with open(script_fm_path + "pointers.json", "r") as file:
            pointers = json.load(file)

        # Check Pointers in json file
        for path in pointers:
            # End the loop if detected new updatas
            if bool_parsing_data == False:
                print("Stop Prasing: Detected New Updates")
                NAS_fm_log("Stop Prasing: Detected New Updates")
                return

            with open(script_fm_path + "pointers.json", "r") as file:
                data = pointers[path]

            if data['Pending to Parse'] == True:
                print(path)
                NAS_fm_log(path)
                print(data['Path'])
                NAS_fm_log(data['Path'])
                file_name = path.replace(data['Path'] + fm_settings.slash,'')
                print(file_name)
                NAS_fm_log(file_name)

                if file_name == (FILENAMES['cp3030']):
                    file_key = 'cp3030'
                elif file_name == (FILENAMES['cp6060']):
                    file_key = 'cp6060'
                elif file_name == (FILENAMES['cv3060']):
                    file_key = 'cv3060'
                elif file_name == (FILENAMES['cv2237']):
                    file_key = 'cv2237'
                elif file_name == (FILENAMES['vis2237']):
                    file_key = 'vis2237'
                elif file_name.__contains__(FILENAMES['prefix']) and file_name.__contains__(FILENAMES['visual']):
                    file_key = 'vis2237' # for other tempertures
                elif file_name == (FILENAMES['focus']):
                    file_key = 'focus'
                else:
                    file_key = 'module_log'
                
                FILENAME = {file_key: file_name}
                # Load Data
                if file_key != 'module_log':
                    try:
                        load_data(data['Path'], FILENAME)
                    except:
                        with open(script_fm_path + "pointers.json","r") as file:
                            pointers = json.load(file)
                            data = pointers[path]

                        # data['Existence'] = False
                        # pointers[path] = data
                        del pointers[path]

                        with open(script_fm_path + "pointers.json","w+") as file:
                            json.dump(pointers, file, indent=4)
                        continue
            

                    global data0
                    df = getattr(data0, file_key)
                    if file_key != 'focus' and 'module_log':
                        Ta = int(file_key[-4:-2])
                        Tb = int(file_key[-2:])
                        df['Ta'] = Ta
                        df['Tb'] = Tb
                    setattr(data0, file_key, df)

                    df = getattr(data0, file_key)
                    if file_key == 'cv3060' or file_key == 'cv2237':
                        try:
                            df['AMBIENT_median'] = df['AMBIENT_median.1']
                            df = df.drop(columns='AMBIENT_median.1')
                        except:
                            pass
                    setattr(data0, file_key, df)

                    if file_key == 'vis2237'and data['Rows'] != data['Pointer']:
                        if file_name == (FILENAMES['vis2237']):
                            data1 = utils.CalibrationData([data['Path']], {'cv2237': (FILENAMES['cv2237'])}, uniquify=False, typify=False, verbose=False)
                            Ta = int('cv2237'[-4:-2])
                            Tb = int('cv2237'[-2:])
                        else:
                            data1 = utils.CalibrationData([data['Path']], {'cv2237': file_name.replace('Visual', 'Verification')}, uniquify=False, typify=False, verbose=False)
                            Ta = int(file_name[-15:-13])
                            Tb = int(file_name[-12:-10])
                        print(Ta)
                        print(Tb)
                        df2 = getattr(data1, 'cv2237')

                        
                        df2['Ta'] = Ta
                        df2['Tb'] = Tb
                        #df['Calibration_Date'] = 'XXXXXXXXXX'
                        setattr(data1, file_key, df2)

                        df2 = getattr(data1, 'cv2237')
                        try:
                            df2['AMBIENT_median'] = df2['AMBIENT_median.1']
                            df2 = df2.drop(columns='AMBIENT_median.1')
                        except:
                            pass
                        try:
                            df2['Delta_Avg'] = df2['SCENE_Avg'] - df2['Tb']
                            df2['Delta_PTAT'] = df2['AMBIENT_median'] - df2['Ta']
                            df2['SCENE_ptp'] = df2['SCENE_hi'] -  df2['SCENE_lo']
                            df2['PTAT_ptp'] = df2['AMBIENT_hi'] -  df2['AMBIENT_lo']
                            df2['Delta_Max'] = df2['SCENE_hi'] - df2['Tb']
                            df2['Delta_Min'] = df2['SCENE_lo'] - df2['Tb']
                            df2['Delta_VDD'] = df2['VDD'] - 3.3
                        except:
                            print("e1")
                        setattr(data1, 'cv2237', df2)


                        #df2 = df2.merge(df[['Serial', 'Original Error', 'Visual Correct Error', 'RESULT_SUMMARY']], how='inner', on='Serial', suffixes=(None, '_vis'))
                        df2 = pd.concat([df2, df['Original Error ' + str(Ta) + str (Tb)]], axis=1)
                        df2 = pd.concat([df2, df['Combine' + str(Ta) + str (Tb) + '_3060_Operator']], axis=1)
                        df2 = pd.concat([df2, df['RESULT_SUMMARY' + str(Ta) + str (Tb)]], axis=1)
                        df = df2.copy()
                        #print(df)

                else:
                    df = pd.read_excel(path)
                    try: # if e-test file has at least one data line
                        df_sfc = get_sfc_frame_module_log(df)
                    except: # if e-test file has no data line
                        with open(script_fm_path + "pointers.json","r") as file:
                            pointers = json.load(file)
                            data = pointers[path]
                        data['Pending to Parse'] = False
                        pointers[path] = data
                        with open(script_fm_path + "pointers.json","w+") as file:
                            json.dump(pointers, file, indent=4)
                            
                        print(f"No Data line in -- {file_key}.")
                        NAS_fm_log(f"No Data line in -- {file_key}.")
                        print()
                        upload_log(data['Path'], "Time:" + get_current_time() + "\tPath: " + path + "\tResult: success\t\tSummary: no data line")
                        NAS_fm_log("Path: " + path + "\tResult: success\t\tSummary: no data line")
                        continue
                if file_key == 'focus':
                    df_sfc = get_sfc_frame_focus(df)
                    #print(file_key)
                if file_key == 'cp3030':
                    df_sfc = get_sfc_frame_calibration_3030_6060(df, "-1")
                    #print(file_key)
                if file_key == 'cp6060':
                    df_sfc = get_sfc_frame_calibration_3030_6060(df, "-2")
                    #print(file_key)
                    pass
                if file_key == 'cv3060':
                    df_sfc = get_sfc_frame_calibration_3060(df)
                    #print(file_key)
                    pass
                if file_key == 'cv2237':
                    #df_sfc = get_sfc_frame_verification_2237(df)
                    #print(file_key)
                    pass
                if file_key == 'vis2237':
                    df_sfc = get_sfc_frame_verification_2237(df)
                    #print(file_key)
                    pass

                sfc_list = df_sfc.to_dict('records')
                line_num = -1
                with requests.Session() as s:
                    success = True
                    for line in sfc_list:
                        print(line)
                        line_num += 1
                        if data['Pointer'] < line_num:
                            try: # skip line if sn number is empty
                                try:
                                    response = s.post(url, json=line, timeout=time_out)
                                    if response.status_code != 201:
                                        print(f'Bad response from server: {response.status_code} for line: {line_num}')
                                        NAS_fm_log(f'Bad response from server: {response.status_code} for line: {line_num}')
                                        upload_log(data['Path'], "Time:" + get_current_time() + "\tPath: " + path + f"\tResult: fail in line {line_num}\t\tSummary: bad response {response.status_code} from server")
                                        success = False
                                        try:
                                            response = s.post(url_error, json=line, timeout=time_out)
                                            data['Pointer'] += 1
                                            update_pointer(path, data['Pointer'])
                                            if fm_settings.bool_print_response:
                                                print(response)
                                                NAS_fm_log(response)
                                            if fm_settings.bool_print_response_content:
                                                print(response.text)
                                                NAS_fm_log(response.text)
                                                print()
                                            continue
                                        except:
                                            return
                                    else:
                                        success = True
                                        data['Pointer'] += 1
                                        update_pointer(path, data['Pointer'])
                                        if fm_settings.bool_print_response:
                                            print(response)
                                            NAS_fm_log(response)
                                        if fm_settings.bool_print_response_content:
                                            print(response.text)
                                            NAS_fm_log(response.text)
                                            print()
                                except requests.exceptions.ConnectionError:
                                    success = False
                                    print(f'Connection Error!')
                                    NAS_fm_log(f'Connection Error!')
                                    print(f"Upload Failed in line: {line_num} -- {file_key}.")
                                    NAS_fm_log(f"Upload Failed in line: {line_num} -- {file_key}.")
                                    upload_log(data['Path'], "Time:" + get_current_time() + "\tPath: " + path + f"\tResult: fail in line {line_num}\t\tSummary: connection error")
                                    NAS_fm_log("Path: " + path + f"\tResult: fail in line {line_num}\t\tSummary: connection error")
                                    return
                                except requests.exceptions.ReadTimeout:
                                    success = False
                                    print(f'Time out!')
                                    NAS_fm_log(f'Time out!')
                                    print(f"Upload Failed in line: {line_num} -- {file_key}.")
                                    NAS_fm_log(f"Upload Failed in line: {line_num} -- {file_key}.")
                                    upload_log(data['Path'], "Time:" + get_current_time() + "\tPath: " + path + f"\tResult: fail in line {line_num}\t\tSummary: time out")
                                    NAS_fm_log("Path: " + path + f"\tResult: fail in line {line_num}\t\tSummary: time out")
                                    return
                            except: # skip line case: if sn number is float('nan')
                                success = True
                                data['Pointer'] += 1
                                update_pointer(path, data['Pointer'])
                                print(f"An empty SN number is detected, skip line {line_num}.")
                                NAS_fm_log(f"An empty SN number is detected, skip line {line_num}.")
                if success:
                    with open(script_fm_path + "pointers.json","r") as file:
                        pointers = json.load(file)
                        data = pointers[path]
                    data['Pending to Parse'] = False
                    pointers[path] = data
                    with open(script_fm_path + "pointers.json","w+") as file:
                        json.dump(pointers, file, indent=4)
                        
                    print(f"Upload Finished -- {file_key}.")
                    NAS_fm_log(f"Upload Finished -- {file_key}.")
                    print()
                    upload_log(data['Path'], "Time:" + get_current_time() + "\tPath: " + path + "\tResult: success\t\tSummary: pass")
                    NAS_fm_log("Path: " + path + "\tResult: success\t\tSummary: pass")
            else:
                print(f"No Updates for {path}.")
                NAS_fm_log(f"No Updates for {path}.")

# the following columns cannot be obtained from the Test_summary_*.tsv or Module_Log_*.xlsx files
# for now we hard code them
add_sfc_columns = {
    #'Ta': 0,
    #'Tb': 0,
    #'qcflag': 0,
    'rework': 0,
    #'station_id': 0 # change in following functions
}

def get_sfc_frame_module_log(df, add_columns=add_sfc_columns):
    try:
        global url 
        url = url_module_log
        df = df.fillna('')
        # station_id, x, y coordinate for module_log
        add_qr_columns = {
            #'station_id': '0',    # station_id
            'mpiID': 0,          # MPI machine id
            'waferCoordX': 0,     # x coordinate
            'waferCoordY': 0,      # y coordinate
            'testType': "ModuleLog"
        }
        sfc2xlsx = {
                # general
            'date': 'Date',
            'serialNumber': 'Serial Number',
            'vendorLot': 'Vendor Lot',
            'operatorID': 'Operator ID',
            'qrCode': 'Hashtag',
            'result': 'Result',
            'current': 'Current',
            'comment':'Comments',
            'moduleType': 'Module Type',
            'batchComment': 'Batch Comment',
            'batchName': 'Batch Name',
            'errorCode': 'Error Code',
            'waferNumber': 'Wafer Number',
        }
        xlsx2sfc = dict([(v, k) for k, v in sfc2xlsx.items()])
        # drop columns if not of interest to SFC
        # this potentially creates a problem if later we need new column name, with
        # value of some columns in the original frame that we have dropped; but ignore for now
        df = df[xlsx2sfc.keys()].copy()
        # change to SFC naming
        df.rename(columns=xlsx2sfc, inplace=True)
        # add x, y coordinate columes
        for k, v in add_qr_columns.items():
            df[k] = v
        # add extra columns, unrelated to original data from the jig
        #for k, v in add_columns.items():
            #df[k] = v
        # special transformations
        df.loc[df.result!='PASS', 'result'] = 0
        df.loc[df.result=='PASS', 'result'] = 1
        df.loc[df.moduleType=='Bobcat with IR04-1', 'moduleType'] = 1
        df.loc[df.moduleType=='Bobcat with IR10-2', 'moduleType'] = 2
        df.loc[df.moduleType=='Bobcat with MI0801-W', 'moduleType'] = 3
        df.loc[df.moduleType=='Bobcat with MI0801-N', 'moduleType'] = 4

        # get station location / x / y coordinate of e-test, by 'serial_number_parser.py'

        for index, item in enumerate(df.serialNumber):
            try:
                args = serial_number_parser.parse_args(item)
                year, week, location, wafer_num, die, x, y = serial_number_parser.serial_parser(args)
            except Exception as e:
                print("serial number is empty")
                print(e)
                wafer_num = location = x = y = None
            df.loc[index, 'mpiID'] = location
            df.loc[index, 'waferCoordX'] = x
            df.loc[index, 'waferCoordY'] = y
            df.loc[index, 'waferNumber'] = wafer_num
            #df.loc[index, 'date'] = datetime.strptime(df.loc[index, 'date'], '%d-%m-%Y %H:%M:%S')
            #print(index)
            # Convert the following from float to int
            #df['mpiID'] = (df['mpiID'].fillna(-1).astype(int).astype(object).where(df['mpiID'].notnull()))
            #df['waferCoordX'] = (df['waferCoordX'].fillna(-1).astype(int).astype(object).where(df['waferCoordX'].notnull()))
            #df['waferCoordY'] = (df['waferCoordY'].fillna(-1).astype(int).astype(object).where(df['waferCoordY'].notnull()))
            #df['waferNumber'] = (df['waferNumber'].fillna(-1).astype(int).astype(object).where(df['waferNumber'].notnull()))

        #print(type(df['date'].iloc[0]))
        #df['date'] = datetime.strptime(df['date'], '%d/%m/%y %H:%M:%S')
        #print(type(df['date'].iloc[0]))
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        df['date'] = df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')
        #print( df['date'])
        #print(type(df['date'].iloc[0]))
        df = df.fillna('')
    except Exception as e:
        print("opps")
        print(e)
    
    return df

def get_sfc_frame_focus(df, add_columns=add_sfc_columns):
    global url 
    url = url_focus
    df = df.fillna('')
    sfc2tsv = {
        # general
        'date': 'Date',
        'serialNumber': 'Serial',
        'vendorLot': 'Vendor lot',
        'operatorID': 'Operator id',
        'stationLocation': 'Station Location',
        'qrCode': 'QRcode',
        'result': 'RESULT_SUMMARY',
        'errorCode': 'Error Code Focus',
        'moduleType': 'Module Type',
        'batchComment': 'Batch Comment',
        'batchName': 'Batch Number',
        # focus station specific
        'ptcValue': 'PTC Value',
        'guiVersion': 'GUI Version',   # PRoduction GUI verison
        'firmwareVersion': 'Firmware Version',     # Jig MCU boards FW
        'position': 'Position',   # Socket position in the jig
        'waferNumber': 'Wafer Number',
        'waferCoordX': 'WaferCoord X',
        'waferCoordY': 'WaferCoord Y',
        'waferErrorCode': 'WaferErrorCode',  # Error code from CP3 wafer-level testing
    }

    tsv2sfc = dict([(v, k) for k, v in sfc2tsv.items()])
    # drop columns if not of interest to SFC
    # this potentially creates a problem if later we need new column name, with
    # value of some columns in the original frame that we have dropped; but ignore for now
    df = df[tsv2sfc.keys()].copy()
    # change to SFC naming
    df.rename(columns=tsv2sfc, inplace=True)

    # add extra columns, unrelated to original data from the jig
    for k, v in add_columns.items():
        df[k] = v

    add_column = {
    'testType': "Focus"
    }

    for k, v in add_column.items():
        df[k] = v

    # print(df.head(50))
    # special transformations
    # print(df.result.values.dtype)
    df.loc[df.result!='FB01:Heat detected', 'result'] = 0 # Fail
    df.loc[df.result=='FB01:Heat detected', 'result'] = 1 # Pass
    df['date'] = df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')
    df = df.fillna('')
    return df


def get_sfc_frame_calibration_3030_6060(df, station_id="-1", add_columns=add_sfc_columns):
    global url 
    url = url_calibration
    #print(url)

    # # get required columns for testdata 
    # testdata_columns = ['PIXEL(40,30)', 'Summary']
    # # merge them into one column 
    # df['testdata_values_list'] = df[testdata_columns].values.tolist()
    df = df.fillna('')
    df['test_result_dict']=df[[
        'LVCM',
        'RVCM',
        'LVDD',
        'RVDD',
        'LPTAT0',
        'LPTAT1',
        'LPTAT2',
        'LPTAT3',
        'RPTAT0',
        'RPTAT1',
        'RPTAT2',
        'RPTAT3',
        'PIXEL(1,2)',
        'PIXEL(78,2)',
        'PIXEL(40,30)',
        'PIXEL(1,63)',
        'PIXEL(78,63)',
        'MedianSTDEV_RawPtat',
        'MedianSTDEV_RawPixel',
        ]].to_dict('records')

    sfc2tsv = {
        # general
        'date': 'Date',
        'serialNumber': 'Serial',
        'vendorLot': 'Vendor lot',
        'operatorID': 'Operator id',
        'result': 'Result',
        'summary': 'Summary',
        'moduleType': 'ModuleType',
        'batchComment': 'Batch Comment',
        'batchName': 'Batch Number',
        'guiVersion': 'GUI Version',   # PRoduction GUI verison
        'firmwareVersion': 'Firmware Version',     # Jig MCU boards FW
        'position': 'Position',   # Socket position in the jig
        'waferNumber': 'Wafer Number',
        'waferCoordX': 'WaferCoord X',
        'waferCoordY': 'WaferCoord Y',
        'waferErrorCode': 'WaferErrorCode',  # Error code from CP3 wafer-level testing
        # 3030/6060 station specific
        'Ta': 'Ta',
        'Tb': 'Tb',
        'testResult': 'test_result_dict',
    }

    tsv2sfc = dict([(v, k) for k, v in sfc2tsv.items()])    
    # drop columns if not of interest to SFC
    # this potentially creates a problem if later we need new column name, with
    # value of some columns in the original frame that we have dropped; but ignore for now
    df = df[tsv2sfc.keys()].copy()
    
    # change to SFC naming
    df.rename(columns=tsv2sfc, inplace=True)
    
    # add extra columns, unrelated to original data from the jig
    for k, v in add_columns.items():
        df[k] = v

    add_column_3030 = {
    'testType': "Calibration3030"
    }
    add_column_6060 = {
    'testType': "Calibration6060"
    }
    if station_id=="-1":
        for k, v in add_column_3030.items():
            df[k] = v
    if station_id=="-2":
        for k, v in add_column_6060.items():
            df[k] = v


    #print(df.head(50))
    # special transformations
    #print(df.result.values.dtype)

    df.loc[df.result=='PASS', 'Summary'] = 'B01' # create B01 upon PASS here
    df.loc[df.result!='PASS', 'result'] = 0 # Fail
    df.loc[df.result=='PASS', 'result'] = 1 # Pass 

    if station_id=="-1":
        df['Ta'] = 30
        df['Tb'] = 30
    if station_id=="-2":
        df['Ta'] = 60
        df['Tb'] = 60
    # df['testConditions']=df[['Ta','Tb']].to_dict('records')
    # df = df.drop('blackbody', 1)
    # df = df.drop('module', 1)

    df['date'] = df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')
    df = df.fillna('')

    return df

def get_sfc_frame_calibration_3060(df, add_columns=add_sfc_columns):
    global url 
    url = url_calibration
    df = df.fillna('')

    # get required columns for testdata 
    df['test_result_dict']=df[[
        'LVCM','RVCM','LVDD','RVDD',
        'LPTAT0','LPTAT1','LPTAT2','LPTAT3',
        'RPTAT0','RPTAT1','RPTAT2','RPTAT3',
        'PIXEL(1,2)','PIXEL(78,2)','PIXEL(40,30)','PIXEL(1,63)','PIXEL(78,63)',
        'PTAT0 off','PTAT0 Grad',
        'PTAT1 off','PTAT1 Grad',
        'PTAT2 off','PTAT2 Grad',
        'PTAT3 off','PTAT3 Grad',
        'PTAT4 off','PTAT4 Grad',
        'PTAT5 off','PTAT5 Grad',
        'PTAT6 off','PTAT6 Grad',
        'PTAT7 off','PTAT7 Grad',
        'Pix[2,1] off','Pix[2,78] off','Pix[30,40] off','Pix[63,1] off','Pix[63,78] off',
        'Pix[2,1] grad','Pix[2,78] grad','Pix[30,40] grad','Pix[63,1] grad','Pix[63,78] grad',
        'PixGrad_ColAvgMax','PixGrad_ColAvgMin',
        'dead pixel','Broken pixel',
        'Pix[2,1]Sen','Pix[2,78]Sen','Pix[30,40]Sen','Pix[63,1]Sen','Pix[63,78]Sen',
        'Pix[2,1]RI','Pix[2,78]RI','Pix[63,1]RI','Pix[63,78]RI',
        'LensShiftX','LensShiftY','VDD','AMBIENT_hi','AMBIENT_lo','SCENE_hi','SCENE_Avg','SCENE_lo',
        'Selected PTAT NR',
        'Fail Pix(B3,B5,B6,B12)',
        'Fail VDD (B10)',
        'Fail PTAT (B10)',
        'Fail PIXGRAD (B10)',
        'Fail/pass Sensitivity (B7 B2 B3)',
        'Fail scene High (B9)',
        'Fail scene low (B9)',
        'Fail Ambient High (B9)',
        'Fail Ambient low (B9)',
        'PTAT Variation',
        'GradVDD_VDD',
        'ADCPermVx100',
        'VDDPixelGradA_Max','VDDPixelGradA_Avg','VDDPixelGradA_Min','VDDPixelGradB_Max','VDDPixelGradB_Avg','VDDPixelGradB_Min',
        'AMBIENT_median',
        'MedianSTDEV_RawPtat',
        'MedianSTDEV_RawPixel',
        ]].to_dict('records')

    sfc2tsv = {
        # general
        'date': 'Date',
        'serialNumber': 'Serial',
        'vendorLot': 'Vendor lot',
        'operatorID': 'Operator id',
        'result': 'RESULT',
        'summary': 'RESULT_SUMMARY',
        'moduleType': 'ModuleType',
        'batchComment': 'Batch Comment',
        'batchName': 'Batch Number',
        'guiVersion': 'GUI Version',   # PRoduction GUI verison
        'firmwareVersion': 'Firmware Version',     # Jig MCU boards FW
        'position': 'Position',   # Socket position in the jig
        'waferNumber': 'Wafer Number',
        'waferCoordX': 'WaferCoord X',
        'waferCoordY': 'WaferCoord Y',
        'waferErrorCode': 'WaferErrorCode',  # Error code from CP3 wafer-level testing
        'stationLocation': 'Station Location',
        # 3060 station specific
        'Ta': 'Ta',
        'Tb': 'Tb',
        'errorCode': 'Error Code Calib',
        'testResult': 'test_result_dict',
    }

    tsv2sfc = dict([(v, k) for k, v in sfc2tsv.items()])    
    # drop columns if not of interest to SFC
    # this potentially creates a problem if later we need new column name, with
    # value of some columns in the original frame that we have dropped; but ignore for now
    df = df[tsv2sfc.keys()].copy()
    
    # change to SFC naming
    df.rename(columns=tsv2sfc, inplace=True)
    
    # add extra columns, unrelated to original data from the jig
    for k, v in add_columns.items():
        df[k] = v

    add_column = {
    'testType': "Sensitivity"
    }

    for k, v in add_column.items():
        df[k] = v

    #print(df.head(50))
    # special transformations
    #print(df.result.values.dtype)

    df.loc[df.result!='PASS', 'result'] = 0 # Fail
    df.loc[df.result=='PASS', 'result'] = 1 # Pass 
    
    
    df['Tb'] = 60
    df['Ta'] = 30
    #df['testConditions']=df[['Tb','Ta']].to_dict('records')

    df['date'] = df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')
    df = df.fillna('')
    return df

def get_sfc_frame_verification_2237(df, add_columns=add_sfc_columns):
    global url 
    url = url_calibration
    df = df.fillna('')
    # get required columns for testdata 
    # get required columns for testdata 
    df['test_result_dict']=df[[
        'LVCM','RVCM','LVDD','RVDD',
        'LPTAT0','LPTAT1','LPTAT2','LPTAT3',
        'RPTAT0','RPTAT1','RPTAT2','RPTAT3',
        'PIXEL(1,2)','PIXEL(78,2)','PIXEL(40,30)','PIXEL(1,63)','PIXEL(78,63)',
        'PTAT0 off','PTAT0 Grad',
        'PTAT1 off','PTAT1 Grad',
        'PTAT2 off','PTAT2 Grad',
        'PTAT3 off','PTAT3 Grad',
        'PTAT4 off','PTAT4 Grad',
        'PTAT5 off','PTAT5 Grad',
        'PTAT6 off','PTAT6 Grad',
        'PTAT7 off','PTAT7 Grad',
        'Pix[2,1] off','Pix[2,78] off','Pix[30,40] off','Pix[63,1] off','Pix[63,78] off',
        'Pix[2,1] grad','Pix[2,78] grad','Pix[30,40] grad','Pix[63,1] grad','Pix[63,78] grad',
        'PixGrad_ColAvgMax','PixGrad_ColAvgMin',
        'dead pixel','Broken pixel',
        'Pix[2,1]Sen','Pix[2,78]Sen','Pix[30,40]Sen','Pix[63,1]Sen','Pix[63,78]Sen',
        'Pix[2,1]RI','Pix[2,78]RI','Pix[63,1]RI','Pix[63,78]RI',
        'LensShiftX','LensShiftY','VDD','AMBIENT_hi','AMBIENT_lo','SCENE_hi','SCENE_Avg','SCENE_lo',
        'Selected PTAT NR',
        'Fail Pix(B3,B5,B6,B12)',
        'Fail VDD (B10)',
        'Fail PTAT (B10)',
        'Fail PIXGRAD (B10)',
        'Fail/pass Sensitivity (B7 B2 B3)',
        'Fail scene High (B9)',
        'Fail scene low (B9)',
        'Fail Ambient High (B9)',
        'Fail Ambient low (B9)',
        'PTAT Variation',
        'GradVDD_VDD',
        'ADCPermVx100',
        'VDDPixelGradA_Max','VDDPixelGradA_Avg','VDDPixelGradA_Min','VDDPixelGradB_Max','VDDPixelGradB_Avg','VDDPixelGradB_Min',
        'AMBIENT_median',
        'MedianSTDEV_RawPtat',
        'MedianSTDEV_RawPixel',
        ]].to_dict('records')

    sfc2tsv = {
        # general
        'date': 'Date',
        'serialNumber': 'Serial',
        'vendorLot': 'Vendor lot',
        'operatorID': 'Operator id',
        'result': 'RESULT',
        'summary': 'RESULT_SUMMARY' + str(df['Ta'][0]) + str(df['Tb'][0]),
        'errorCode': 'Error Code Calib',
        'qrCode': 'QRcode',
        'moduleType': 'ModuleType',
        'batchComment': 'Batch Comment',
        'batchName': 'Batch Number',
        'guiVersion': 'GUI Version',   # PRoduction GUI verison
        'firmwareVersion': 'Firmware Version',     # Jig MCU boards FW
        'position': 'Position',   # Socket position in the jig
        'waferNumber': 'Wafer Number',
        'waferCoordX': 'WaferCoord X',
        'waferCoordY': 'WaferCoord Y',
        'waferErrorCode': 'WaferErrorCode',  # Error code from CP3 wafer-level testing
        'stationLocation': 'Station Location',
        # 2237 station specific
        'Ta': 'Ta',
        'Tb': 'Tb',
        'originalError': 'Original Error ' + str(df['Ta'][0]) + str(df['Tb'][0]),
        'visualCorrectError': 'Combine'+ str(df['Ta'][0]) + str(df['Tb'][0]) + '_3060_Operator',
        'testResult': 'test_result_dict',
    }

    tsv2sfc = dict([(v, k) for k, v in sfc2tsv.items()])    
    # drop columns if not of interest to SFC
    # this potentially creates a problem if later we need new column name, with
    # value of some columns in the original frame that we have dropped; but ignore for now
    df = df[tsv2sfc.keys()].copy()
    
    # change to SFC naming
    df.rename(columns=tsv2sfc, inplace=True)
    
    # add extra columns, unrelated to original data from the jig
    for k, v in add_columns.items():
        df[k] = v

    add_column = {
    'testType': "Verification"
    }

    for k, v in add_column.items():
        df[k] = v

    #print(df.head(50))
    # special transformations
    #print(df.result.values.dtype)

    df.loc[df.errorCode!='B00', 'result'] = 0 # Fail
    df.loc[df.errorCode=='B00', 'result'] = 1 # Pass 
    
    #df['Tb'] = 37
    #df['Ta'] = 22
    #df['testConditions']=df[['Tb','Ta']].to_dict('records')

    df['date'] = df['date'].dt.strftime('%Y-%m-%d %H:%M:%S')
    df = df.fillna('')
    #print(df)
    return df

def check_pointer(filepath):
    #Check Pointer
    global bool_changes
    with open(script_fm_path + "pointers.json", "r") as file:
        pointers = json.load(file)

    if filepath in pointers:
        update_pointer(filepath, -1)
    else:
        create_pointer(filepath)
        bool_changes = True

def scanning_file(path):
    cp3030 = FILENAMES['cp3030']
    cp6060 = FILENAMES['cp6060']
    cv3060 = FILENAMES['cv3060']
    cv2237 = FILENAMES['cv2237']
    vis2237 = FILENAMES['vis2237']
    focus = FILENAMES['focus']
    prefix = FILENAMES['prefix']
    visual = FILENAMES['visual']
    module_log = 'Module_Log'
    for filename in os.listdir(path):
        if filename.__contains__(cp3030): 
            filepath = path + r'/' + cp3030
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)
            
        if filename.__contains__(cp6060): 
            filepath = path + r'/' + cp6060
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)

        if filename.__contains__(cv3060): 
            filepath = path + r'/' + cv3060
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)

        # if filename.__contains__(cv2237): 
        #     filepath = path + r'/' + cv2237
        #     print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
        #     NAS_fm_log(" on_modified " + filepath)
        #     check_pointer(filepath)

        if filename.__contains__(vis2237): 
            filepath = path + r'/' + vis2237
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)

        if filename.__contains__(prefix) and filename.__contains__(visual) and filename != vis2237: 
            filepath = path + r'/' + filename
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)

        if filename.__contains__(focus): 
            filepath = path + r'/' + focus
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)

        if filename.__contains__(module_log): 
            filepath = path + r'/' + filename
            print(time.strftime("\r%Y-%m-%d %H:%M:%S", time.gmtime()) + " on_modified " + filepath)
            NAS_fm_log(" on_modified " + filepath)
            check_pointer(filepath)

def e_test_new_folder(path):
    global new_folder_count
    i = 1
    try:
        print(f"{path} cannot be found")
        NAS_fm_log(f"{path} cannot be found")
        directory = os.path.dirname(path)
        while  i < new_folder_count:
            directory = os.path.dirname(directory)
            i = i + 1
        path = max([os.path.join(directory,d) for d in os.listdir(directory)], key=os.path.getmtime)
        i = 1
        while  i < new_folder_count:
            path = max([os.path.join(path,d) for d in os.listdir(path)], key=os.path.getmtime)
            i = i + 1

        #NAS_fm_log(" on_modified " + filepath)
        print("Checking for: " + path)
        NAS_fm_log("Checking for: " + path)
        scanning_file(path)
    except:
        new_folder_count = new_folder_count + 1
        e_test_new_folder(path)


class MyHandler(FileSystemEventHandler):
    def on_modified(self, event):
        path = (event.src_path).replace("\\", r"/")
        print(path)

        if ((path.__contains__('E-Test') or path.__contains__('CalibStation') or path.__contains__('FocusStation') or path.__contains__('FocusCheckStation')) and not (path.__contains__('Event_Log') or path.__contains__('Temp_Log') or path.__contains__('.txt') or path.__contains__('.bmp') or path.__contains__('.bin') or path.__contains__('PictureBox') or path.__contains__('rawdata'))):
            global bool_parsing_data
            bool_parsing_data = False

            #os.system('cls' if os.name == 'nt' else 'clear')
            try:
                scanning_file(path)
            except:
                print(f"cannot found {path}")
                NAS_fm_log(f"cannot found {path}")
                e_test_new_folder(path)          
            
            if bool_changes:
                t = Timer(0.1, fm_main)
                t.start()
            

event_handler = MyHandler()
observer = Observer()
observer.schedule(event_handler, folder_to_track, recursive = True)
observer.start()

def daily():
    while 1==1:
        if time.strftime("%H") == "11": # Runs every night at 11pm
            new_day()
            delete_outdated_pointer()
        time.sleep(40*60) # Check the time every 40 mins
        
print("for testing")
print("============== Start NAS folder monitor ==============")
NAS_fm_log("Start NAS folder monitor")

day = Timer(0, daily)
day.start()

try:
    while True:
        time.sleep(1)
except (KeyboardInterrupt, SystemExit):
    observer.stop()

observer.join()
